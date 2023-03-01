import datetime
import openpyxl
from asgiref.sync import sync_to_async


async def save_order_data_to_excel(order_data: dict):
    now = datetime.datetime.now()
    filename = "orders.xlsx"

    try:
        wb = await sync_to_async(openpyxl.load_workbook)(filename)
    except FileNotFoundError:
        wb = await sync_to_async(openpyxl.Workbook)()

    ws = wb.active

    if not ws.cell(1, 1).value:
        ws.cell(1, 1, "Username")
        ws.cell(1, 2, "Products")
        ws.cell(1, 3, "Payment amount")
        ws.cell(1, 4, "Payment ID")
        ws.cell(1, 5, "Order date")
        ws.cell(1, 6, "address")

    next_row = ws.max_row + 1

    ws.cell(next_row, 1, order_data["username"])

    products = [f"{product[0]} {product[1]} шт." for product in order_data["products"]]
    ws.cell(next_row, 2, "\n".join(products))

    ws.cell(next_row, 3, order_data["payment_amount"])
    ws.cell(next_row, 4, order_data["payment_id"])
    ws.cell(next_row, 5, now.strftime("%Y-%m-%d %H:%M:%S"))
    ws.cell(next_row, 6, order_data["address"])

    await sync_to_async(wb.save)(filename)
